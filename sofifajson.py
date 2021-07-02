import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import openpyxl.styles as sty
from openpyxl import Workbook ,load_workbook
import json

wb=load_workbook(".\\sofifa.xlsx")#打开sofifa.xlsx
ws = wb["Sheet1"]#打开xlsx的sheet1
list_key=[]#用于存放键值名称
jsonLine=[]#用于存放球员数据
for col in range(1,4+1):
    list_key.append(ws.cell(row=1,column=col).value)#生成键值名称
for row in range(2,3601+1):#对行循环
    dict_v={}#用来存放该循环球员的json格式的信息
    for col in range(1,1+4):#添加json格式的信息
        dict_v[list_key[col-1]]=ws.cell(row=row,column=col).value#将该球员信息添加到jsonLine中
    jsonLine.append(dict_v)#将获取信息append到jsonLine中
    
json.dump(jsonLine,open(".\\sofifa.txt","w"),indent=0)#dump函数将json写入sofifa.txt